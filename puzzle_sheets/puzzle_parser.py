import logging
import os

from collections import OrderedDict
from collections import namedtuple
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side
from openpyxl.styles.fonts import DEFAULT_FONT
from openpyxl.utils.cell import get_column_letter
from puz import Puzzle

import puz


Solution = namedtuple('Solution', ('num', 'clue', 'len', 'answer'))

# Style Settings
DEFAULT_FONT.name = 'Arial'
DEFAULT_FONT.size = 10
BOLD_FONT = Font(name='Arial', size=10, bold=True)
SOLID_BLOCK = PatternFill('solid', bgColor='000000')
SOLID_SIDE = Side(border_style='thin', color='000000')
SOLID_BORDER = border = Border(left=SOLID_SIDE, top=SOLID_SIDE, right=SOLID_SIDE, bottom=SOLID_SIDE)

# Based on trial and error. ¯\_(ツ)_/¯
CLUE_WIDTH = 60
GRID_WIDTH = 3


def _get_across_answer(puzzle: Puzzle, clue: dict) -> str:
    """Returns the across answer for a given clue in the puzzle. Any blanks will be indicated by dashes."""
    return puzzle.fill[clue['cell'] : clue['cell'] + clue['len']]


def _get_down_answer(puzzle: Puzzle, clue: dict) -> str:
    """Returns the down answer for a given clue in the puzzle. Any blanks will be indicated by dashes."""
    return ''.join([puzzle.fill[clue['cell'] + i * puzzle.width] for i in range(clue['len'])])


def _get_solution_grid(puzzle: Puzzle) -> dict:
    """Returns a dict that contains the whole grid. The dict has 2 keys: across and down. The value of each is
    an OrderedDict whose key is the clue number, and value is a Solution namedtuple."""
    numbering = puzzle.clue_numbering()
    return {
        'across': OrderedDict(
            {
                clue['num']: Solution(clue['num'], clue['clue'], clue['len'], _get_across_answer(puzzle, clue))
                for clue in numbering.across
            }
        ),
        'down': OrderedDict(
            {
                clue['num']: Solution(clue['num'], clue['clue'], clue['len'], _get_down_answer(puzzle, clue))
                for clue in numbering.down
            }
        ),
    }


def _get_clue_str(solution: Solution) -> str:
    """Returns a string representation of the clue, for printing."""
    return f'{solution.num}. {solution.clue} ({solution.len})'


def validate(puz_path: str) -> Puzzle:
    """Validate the input file, and return contents if valid"""
    if not puz_path.upper().endswith('.PUZ'):
        logging.error('Only .PUZ files are supported.')
    elif not os.path.isfile(puz_path):
        logging.error('Cannot read file %s', puz_path)
    else:
        try:
            return puz.read(puz_path)
        except puz.PuzzleFormatError:
            logging.exception('Cannot validate file %s', puz_path)


def display(puzzle: Puzzle):
    """Prints PUZ file contents to terminal."""
    grid = _get_solution_grid(puzzle)

    print('ACROSS')
    for solution in grid['across'].values():
        print(_get_clue_str(solution))

    print('DOWN')
    for solution in grid['down'].values():
        print(_get_clue_str(solution))

    for row in range(puzzle.height):
        cell = row * puzzle.width
        print('\t'.join(puzzle.fill[cell : cell + puzzle.width]).replace('.', chr(9608) * 2))

    print(f'Title: {puzzle.title}')
    print(f'Author: {puzzle.author}')
    print(f'Note: {puzzle.notes}')


def _write_metadata(sheet, row, key, value):
    """Write the metadata of the puzzle, below the grid."""
    key_cell = _get_cell(sheet, row, 1)
    key_cell.value = key
    key_cell.font = BOLD_FONT
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)

    value_cell = _get_cell(sheet, row, 3)
    value_cell.value = value or 'N.A.'


def _get_cell(sheet, row, col):
    return sheet.cell(row=row, column=col)


def _get_xlsx_path(puz_path: str) -> str:
    if puz_path.endswith('.puz'):
        xlsx_path = puz_path.removesuffix('.puz')
    else:
        xlsx_path = puz_path.removesuffix('.PUZ')
    tstamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    return f'{xlsx_path}_{tstamp}.xlsx'


def write_xlsx(puz_path: str, puzzle: Puzzle) -> str:
    """Writes the PUZ file contents to a XLSX file and returns its filename."""
    grid = _get_solution_grid(puzzle)
    workbook = Workbook()
    sheet = workbook.active

    # Across clues
    across_start_col = puzzle.width + 1
    cell = _get_cell(sheet, 1, across_start_col)
    cell.value = 'ACROSS'
    cell.font = BOLD_FONT
    sheet.column_dimensions[get_column_letter(across_start_col)].width = CLUE_WIDTH
    for index, solution in enumerate(grid['across'].values()):
        _get_cell(sheet, index + 2, across_start_col).value = _get_clue_str(solution)

    # Down clues
    down_start_col = across_start_col + 1
    cell = _get_cell(sheet, 1, down_start_col)
    cell.value = 'DOWN'
    cell.font = BOLD_FONT
    sheet.column_dimensions[get_column_letter(down_start_col)].width = CLUE_WIDTH
    for index, solution in enumerate(grid['down'].values()):
        _get_cell(sheet, index + 2, down_start_col).value = _get_clue_str(solution)

    # Solution
    for row in range(puzzle.height):
        cell_index = row * puzzle.width
        for index, char in enumerate(puzzle.fill[cell_index : cell_index + puzzle.width]):
            cell = sheet.cell(row=row + 1, column=index + 1)
            cell.border = SOLID_BORDER
            if char == '.':
                cell.fill = SOLID_BLOCK
            else:
                cell.value = char

    # Metadata
    metadata_start_row = puzzle.height + 1
    _write_metadata(sheet, metadata_start_row, 'Title', puzzle.title)
    _write_metadata(sheet, metadata_start_row + 1, 'Author', puzzle.author)
    _write_metadata(sheet, metadata_start_row + 2, 'Note', puzzle.notes)

    # Styling
    for col in range(puzzle.width):
        sheet.column_dimensions[get_column_letter(col + 1)].width = GRID_WIDTH

    xlsx_path = _get_xlsx_path(puz_path)
    logging.info('Writing XLSX %s from PUZ %s', xlsx_path, puz_path)
    workbook.save(xlsx_path)
    return xlsx_path
